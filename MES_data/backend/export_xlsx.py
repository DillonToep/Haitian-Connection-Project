"""Builds the "试模成型参数表" (trial-mold parameter form) as an .xlsx
file, filling in whichever cells have a matching value already stored in
the MES. Two code paths share the same cell mapping
(PARAMETER_CELL_MAP / EXTENDED_CELL_MAP / HEADER_CELL_MAP):

- build_trial_parameter_workbook(): generates a brand-new workbook from
  the company's original paper/.xls template (grid text + merged cells
  captured verbatim below as static data, extracted once from
  参数.xls via xlrd, so this module has no runtime dependency on that
  file). Used when no workbook has ever been uploaded for a given Mold +
  Machine Type -- see backend/template_storage.py.

- overlay_values_onto_template(): writes the same mapped values onto a
  COPY of a workbook a user previously uploaded (see
  backend/routers/export.py's POST .../import), instead of regenerating
  a sheet. Every cell not in the mapping -- and all formatting, merges,
  images, formulas, column widths, page setup, etc. -- is left exactly
  as uploaded. This is the path used once a Mold + Machine Type has an
  uploaded template on file, so "upload -> edit in app -> export" writes
  back into the user's own file rather than a freshly generated one.

Only cells with a confident, unambiguous field mapping are filled in --
see PARAMETER_CELL_MAP / EXTENDED_CELL_MAP / HEADER_CELL_MAP below. Fields
that only exist on the paper form (试模员/试模日期/审核/试模结果 defect
checklist, 螺杆直径, 锁模力, etc.) are intentionally left blank since the
MES has no data behind them -- the exported sheet is meant to be finished
by hand during the actual trial run.
"""
import json
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------
# Static template data: {"rows": [[str,...], ...], "merges": [[r0,r1,c0,c1], ...]}
# rows/merges use the same 0-indexed, half-open convention xlrd returns
# (a merge covers rows r0..r1-1 and cols c0..c1-1).
# ---------------------------------------------------------------------
_TEMPLATE = json.loads(r'''{"rows": [["试模成型参数表", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""], ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""], ["乔丰科技实业（深圳）有限公司", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""], ["QiaoFeng  technology industrial (shenzhen) co., LTD", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""], ["模具编号", "", "", "产品名称", " ", "", "模穴数", "", "试模次数", "", "", "试模日期", "", "", "", "生产啤数", ""], ["胶料名称", "", "", "   客供   本厂料", "", "颜色", "", "色粉编号", "", "", "烘料温度", "", "烘料方式：□普通  □抽湿", "℃", "", "烘料时间：     H", "   H"], ["试模员", "", "", "要求上机时间", "", "", "", "", "实际完成时间", "", "", "", "", "", "试模用料重量  kg", "", "   Kg"], ["机位/机型", "", "", "品牌", "", "螺杆直径：  MM", "", "安数：      Oz", "", "", " 锁模力 ：         KN       ", "", "", "", "最大注塑压力      ", "        kgf/cm2", ""], ["包装要求", " 普通   出口", "样板要求", "     卡尺寸    客看   项目    自确认", "", "", "", "", "", "", "净重     g", " ", "", "毛重", " ", "1啤总重量", ""], ["料温设定温度\\n（℃）", "射嘴", "1段", "2 段", "3段", "4段", "5段", "热流道设定温度（℃）", "热流板", "", "1段", "2 段", "3段", "3段", "4段", "5段", "6段"], ["", "", "", "", "", "", "", "", " ", "", "", "", "", "", "", "", ""], ["模温设定\\n（℃）", "前模", "后模", "行位", "实测模温", "前模", "后模", "行位", "啤试模式", "", "  全自动   半自动    手动    机械手", "", "", "", "", "", ""], ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""], ["射胶设定", "", "1级", "2级", "3级", "4级", "5级", "射胶时间（秒）", "冷却时间（S)", "", "保压切换位置（MM)", "保压", "", "", "1级", "2级", "3级"], ["", "速度", "", "", "", "", "", "", "", "", "", "", "", "速度", "", "", ""], ["", "压力", "", "", "", "", "", "", "", "", "", "", "", "压力", "", "", ""], ["", "位置", "", "", "", "", "", "", "", "", "", "", "", "时间（S)", "", "", ""], ["锁模设定±10℅", "", "1段", "2段", "3段", "高压", "开模设定±10℅", "", "1段", "2段", "2段", "3段", "", "4段", "顶出方式", "顶出次数", "顶出行程（MM)"], ["", "速度", "", "", "", "", "", "速度", "", "", "", "", "", "", "   拉杆", "", ""], ["", "压力", "", "", "", "", "", "压力", "", "", "", "", "", "", " 普通", "", ""], ["", "位置", "", "", "", "", "", "位置", "", "", "", "", "", "", " 油缸", "", ""], ["熔胶设定", "", "1段", "2段", "3段", "4段", "螺杆转速（RPM）", "", "熔胶时间（S）", "", "", "松退（MM）", "", "", "残余料量（MM）", "", "周期（S）"], ["", "速度", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""], ["", "压力", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""], ["", "位置", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""], ["", "背压", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""], ["试模结果", "", "", "", "", "", "", "", "其他", "", "", "", "", "", "", "", ""], ["  分型面披峰/抹模", "", "", "", "□模花", "", "", "", "", "", "", "", "", "", "", "", ""], ["□行位披峰", "", "", "", "□顶白/拉白", "", "", "", "", "", "", "", "", "", "", "", ""], ["□孔位披峰", "", "", "", "  走胶不齐", "", "", "", "", "", "", "", "", "", "", "", ""], ["□柱位披峰", "", "", "", "□顶烂", "", "", "", "", "", "", "", "", "", "", "", ""], ["  斜顶位披峰", "", "", "", "□断顶针/司筒/螺丝", "", "", "", "", "", "", "", "", "", "", "", ""], ["□分型面夹口不平", "", "", "", "□模点", "", "", "", "", "", "", "", "", "", "", "", ""], ["□行位夹口不平", "", "", "", "  困气/烧胶", "", "", "", "", "", "", "", "", "", "", "", ""], ["□斜顶夹口不平", "", "", "", "□气泡", "", "", "", "", "", "", "", "", "", "", "", ""], ["□枕位夹口不平", "", "", "", "  气纹/夹水纹", "", "", "", "", "", "", "", "", "", "", "", ""], ["□顶针位披峰", "", "", "", "□变形/曲翘", "", "", "", "", "", "", "", "", "", "", "", ""], ["□顶针位下陷不平", "", "", "", "□油污清洗不干净", "", "", "", "", "", "", "", "", "", "", "", ""], ["□顶针位高出平面", "", "", "", "□A板漏水", "", "", "", "", "", "", "", "", "", "", "", ""], ["  粘水口", "", "", "", "  B板漏水", "", "", "", "", "", "", "", "", "", "", "", ""], ["  粘A板", "", "", "", "□行位漏水", "", "", "", "", "", "", "", "", "", "", "", ""], ["□粘B板", "", "", "", "□A板运水不通", "", "", "", "", "", "", "", "", "", "", "", ""], ["□粘行位", "", "", "", "□B板运水不通", "", "", "", "", "", "", "", "", "", "", "", ""], ["   拖花/拉伤", "", "", "", "□行位运水不通", "", "", "", "", "", "", "", "", "", "", "", ""], ["□塌边/边凹凸不平", "", "", "", "□行位不顺/行不到位", "", "", "", "", "", "", "", "", "", "", "", ""], ["  骨位火花纹/线割纹未省光", "", "", "", "□顶针不顺/顶不动/退不回", "", "", "", "", "", "", "", "", "", "", "", ""], ["□刀纹/火花纹未省模", "", "", "", "□五金装不到位", "", "", "", "", "", "", "", "", "", "", "", ""], ["表格编号：QF-QR-009B-057-E2", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""], ["审核：", "", "", "日期： ", "", "", "", "", "", "", "", "", "", "", "", "", ""]], "merges": [[2, 3, 0, 17], [3, 4, 0, 17], [4, 5, 1, 3], [4, 5, 4, 6], [4, 5, 13, 15], [5, 6, 1, 3], [5, 6, 3, 5], [5, 6, 10, 12], [5, 6, 13, 15], [6, 7, 1, 3], [6, 7, 3, 5], [6, 7, 5, 8], [6, 7, 8, 11], [6, 7, 11, 14], [7, 8, 1, 3], [7, 8, 15, 17], [8, 9, 3, 9], [10, 11, 12, 14], [26, 27, 0, 8], [26, 27, 8, 17], [27, 28, 0, 4], [27, 28, 4, 8], [28, 29, 0, 4], [28, 29, 4, 8], [29, 30, 0, 4], [29, 30, 4, 8], [30, 31, 0, 4], [30, 31, 4, 8], [31, 32, 0, 4], [31, 32, 4, 8], [32, 33, 0, 4], [32, 33, 4, 8], [33, 34, 0, 4], [33, 34, 4, 8], [34, 35, 0, 4], [34, 35, 4, 8], [35, 36, 0, 4], [35, 36, 4, 8], [36, 37, 0, 4], [36, 37, 4, 8], [37, 38, 0, 4], [37, 38, 4, 8], [38, 39, 0, 4], [38, 39, 4, 8], [39, 40, 0, 4], [39, 40, 4, 8], [40, 41, 0, 4], [40, 41, 4, 8], [41, 42, 0, 4], [41, 42, 4, 8], [42, 43, 0, 4], [42, 43, 4, 8], [43, 44, 0, 4], [43, 44, 4, 8], [44, 45, 0, 4], [44, 45, 4, 8], [45, 46, 0, 4], [45, 46, 4, 8], [46, 47, 0, 4], [46, 47, 4, 8], [9, 11, 0, 1], [11, 13, 0, 1], [13, 17, 0, 1], [17, 21, 0, 1], [21, 26, 0, 1], [11, 13, 4, 5], [17, 21, 6, 7], [9, 11, 7, 8], [13, 15, 7, 8], [15, 17, 7, 8], [11, 13, 8, 9], [13, 15, 8, 9], [15, 17, 8, 9], [13, 15, 10, 11], [15, 17, 10, 11], [13, 17, 11, 12], [17, 19, 15, 16], [19, 21, 15, 16], [17, 19, 16, 17], [19, 21, 16, 17], [21, 24, 16, 17], [24, 26, 16, 17], [24, 26, 8, 11], [24, 26, 11, 14], [21, 24, 6, 8], [21, 24, 14, 16], [24, 26, 6, 8], [24, 26, 14, 16], [21, 24, 8, 11], [21, 24, 11, 14], [11, 13, 10, 17], [27, 47, 8, 17], [0, 2, 0, 16]]}''')

_ROWS = _TEMPLATE["rows"]
_MERGES = _TEMPLATE["merges"]

_THIN = Side(style="thin", color="444B57")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_LABEL_FONT = Font(name="Arial", size=9)
_TITLE_FONT = Font(name="Arial", size=16, bold=True)
_SUBTITLE_FONT = Font(name="Arial", size=11, bold=True)
_VALUE_FONT = Font(name="Arial", size=10, bold=True, color="1F4A35")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ---------------------------------------------------------------------
# Field mappings: (0-indexed row, 0-indexed col) -> lookup key.
# "param:<TAG>" pulls from the 高级工艺参数 target_value for that tag.
# "ext:<key>" pulls from the mold's extended-info fields JSON.
# "mold:<key>" pulls directly from the mold record.
# Only unambiguous, high-confidence matches are listed -- see module
# docstring for what was deliberately left out.
# ---------------------------------------------------------------------

HEADER_CELL_MAP = {
    (4, 1): "mold:mold_code",
    (4, 4): "mold:mold_name",
    (4, 7): "mold:cavities",
}

EXTENDED_CELL_MAP = {
    (5, 5): "ext:color",
    (5, 8): "ext:color_code",
    (5, 11): "ext:oven_temperature",
    (7, 4): "ext:machine_maker",
    (8, 11): "ext:net_weight",
    (8, 14): "ext:gross_weight",
}

# Hot-runner temperature stages -- template has 7 slots (row9 cols10-16,
# data row10 cols10-16) matching hot_runner_t1..t7.
_HOT_RUNNER_ROW = 10
_HOT_RUNNER_COLS = [10, 11, 12, 13, 14, 15, 16]

PARAMETER_CELL_MAP = {
    # 料温设定温度 -- 1段..5段 (row10, cols2-6). 射嘴/nozzle (col1) has no
    # backing 工艺参数 tag (frontend keeps it as a local-only value), so
    # it's left blank.
    (10, 2): "param:TS1",
    (10, 3): "param:TS2",
    (10, 4): "param:TS3",
    (10, 5): "param:TS4",
    (10, 6): "param:TS5",

    # 射胶设定 速度/压力/位置, 1级..5级 (rows 14/15/16, cols2-6)
    (14, 2): "param:IV1", (14, 3): "param:IV2", (14, 4): "param:IV3", (14, 5): "param:IV4", (14, 6): "param:IV5",
    (15, 2): "param:IP1", (15, 3): "param:IP2", (15, 4): "param:IP3", (15, 5): "param:IP4", (15, 6): "param:IP5",
    (16, 2): "param:IS1", (16, 3): "param:IS2", (16, 4): "param:IS3", (16, 5): "param:IS4", (16, 6): "param:IS5",

    # 保压 速度/压力/时间, 1级..3级 (rows 14/15/16, cols14-16)
    (14, 14): "param:PV1", (14, 15): "param:PV2", (14, 16): "param:PV3",
    (15, 14): "param:PP1", (15, 15): "param:PP2", (15, 16): "param:PP3",
    (16, 14): "param:PT1", (16, 15): "param:PT2", (16, 16): "param:PT3",

    # 锁模设定 速度/压力/位置, 1段/2段/3段/高压 (rows18/19/20, cols2-5) --
    # the sheet only has 4 lock-clamp columns while the MES tracks 5
    # stages (MCV1-5); "高压" (high-pressure/final stage) maps to
    # stage 5, matching how the machine's own panel labels its last
    # lock-clamp stage.
    (18, 2): "param:MCV1", (18, 3): "param:MCV2", (18, 4): "param:MCV3", (18, 5): "param:MCV5",
    (19, 2): "param:MCP1", (19, 3): "param:MCP2", (19, 4): "param:MCP3", (19, 5): "param:MCP5",
    (20, 2): "param:MCS1", (20, 3): "param:MCS2", (20, 4): "param:MCS3", (20, 5): "param:MCS5",

    # 开模设定 速度/压力/位置, 1段..4段 (rows18/19/20, cols8,9,11,13 --
    # col10 is a duplicate "2段" header in the original sheet and is
    # skipped)
    (18, 8): "param:MOV1", (18, 9): "param:MOV2", (18, 11): "param:MOV3", (18, 13): "param:MOV4",
    (19, 8): "param:MOP1", (19, 9): "param:MOP2", (19, 11): "param:MOP3", (19, 13): "param:MOP4",
    (20, 8): "param:MOS1", (20, 9): "param:MOS2", (20, 11): "param:MOS3", (20, 13): "param:MOS4",

    # 顶出次数 (row18, col15)
    (18, 15): "param:EJET",

    # 熔胶设定 速度/压力/位置/背压, 1段..4段 (rows22-25, cols2-5)
    (22, 2): "param:PLV1", (22, 3): "param:PLV2", (22, 4): "param:PLV3", (22, 5): "param:PLV4",
    (23, 2): "param:PLP1", (23, 3): "param:PLP2", (23, 4): "param:PLP3", (23, 5): "param:PLP4",
    (24, 2): "param:PLS1", (24, 3): "param:PLS2", (24, 4): "param:PLS3", (24, 5): "param:PLS4",
    (25, 2): "param:PLBP1", (25, 3): "param:PLBP2", (25, 4): "param:PLBP3", (25, 5): "param:PLBP4",

    # 熔胶时间(S) / 周期(S) -- single-value fields with a confident name
    # match (rows22, cols8 and 16)
    (22, 8): "param:EPLST",
    (22, 16): "param:ECYCT",

    # 冷却时间(S) (row10... actually its own header row13 col8, data
    # lives in the merged block starting row14 col8)
    (14, 8): "param:CT",
}


def _tag_value(parameters_by_tag: dict, tag: str):
    row = parameters_by_tag.get(tag)
    if row is None:
        return None
    value = row.get("target_value") if isinstance(row, dict) else None
    if value in (None, ""):
        return None
    return value


def _extended_value(extended_fields: dict, key: str):
    value = extended_fields.get(key)
    if value in (None, ""):
        return None
    return value


def build_trial_parameter_workbook(mold: dict, parameters_by_tag: dict, extended_fields: dict) -> BytesIO:
    """mold: {"mold_code","mold_name","cavities"}.
    parameters_by_tag: {parameter_id: {"target_value": ...}} for one
    Mold + Machine Type (same rows dbo.mold_parameter_targets holds).
    extended_fields: the machine type's 高级参数 extended-info fields dict
    (dbo.mold_extended_info.info_json, already parsed).
    Returns an in-memory .xlsx file (BytesIO, position 0)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "试模成型参数表"
    ws.sheet_view.showGridLines = False

    for col_idx in range(1, 18):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10.5

    # ---- reproduce the original grid text + borders ----
    for r0, row_values in enumerate(_ROWS):
        for c0, text in enumerate(row_values):
            cell = ws.cell(row=r0 + 1, column=c0 + 1)
            if text not in (None, ""):
                cell.value = text
            cell.border = _BORDER
            cell.alignment = _CENTER
            cell.font = _LABEL_FONT

    ws.cell(row=1, column=1).font = _TITLE_FONT
    ws.cell(row=3, column=1).font = _SUBTITLE_FONT

    # ---- reproduce merges ----
    for r0, r1, c0, c1 in _MERGES:
        ws.merge_cells(start_row=r0 + 1, start_column=c0 + 1, end_row=r1, end_column=c1)

    # Any target cell that falls inside a merged range must be written via
    # its top-left anchor -- openpyxl's other cells in a merge are
    # read-only MergedCell placeholders. Build a (row0,col0) -> anchor
    # lookup once so callers below can just name the cell they mean.
    _anchor_for = {}
    for r0, r1, c0, c1 in _MERGES:
        for rr in range(r0, r1):
            for cc in range(c0, c1):
                _anchor_for[(rr, cc)] = (r0, c0)

    # ---- overlay values (only where the MES actually has data) ----
    def _set(row0, col0, value):
        if value is None:
            return
        anchor_row0, anchor_col0 = _anchor_for.get((row0, col0), (row0, col0))
        cell = ws.cell(row=anchor_row0 + 1, column=anchor_col0 + 1)
        cell.value = value
        cell.font = _VALUE_FONT

    for (r0, c0), key in HEADER_CELL_MAP.items():
        _, field = key.split(":", 1)
        _set(r0, c0, mold.get(field))

    for (r0, c0), key in EXTENDED_CELL_MAP.items():
        _, field = key.split(":", 1)
        _set(r0, c0, _extended_value(extended_fields, field))

    for (r0, c0), key in PARAMETER_CELL_MAP.items():
        _, tag = key.split(":", 1)
        _set(r0, c0, _tag_value(parameters_by_tag, tag))

    for offset, col0 in enumerate(_HOT_RUNNER_COLS):
        tag_key = f"ext:hot_runner_t{offset + 1}"
        _, field = tag_key.split(":", 1)
        _set(_HOT_RUNNER_ROW, col0, _extended_value(extended_fields, field))

    ws.row_dimensions[1].height = 26
    ws.freeze_panes = None

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def overlay_values_onto_template(
    file_bytes: bytes,
    is_macro_enabled: bool,
    mold: dict,
    parameters_by_tag: dict,
    extended_fields: dict,
) -> BytesIO:
    """Writes the same mapped values as build_trial_parameter_workbook,
    but onto a COPY of a user-uploaded workbook (file_bytes) instead of
    regenerating a sheet from the embedded _TEMPLATE. Only the cells
    listed in HEADER_CELL_MAP / EXTENDED_CELL_MAP / PARAMETER_CELL_MAP /
    the hot-runner columns are touched -- every other cell, its
    formatting, merges, images, formulas, column widths, page setup,
    etc. are left exactly as uploaded.

    Unlike build_trial_parameter_workbook (which skips None values,
    since it starts from an already-blank template), a mapped field with
    no value here explicitly clears its cell to blank -- the uploaded
    workbook may already hold a stale value from a previous edit/export
    cycle, and per spec a field with no value must blank its cell rather
    than silently keep whatever was there before.

    Merge ranges are read from the ACTUAL uploaded worksheet (not the
    static _MERGES table used by build_trial_parameter_workbook), so this
    works correctly even if the uploaded file's structure isn't byte-for-
    byte identical to the embedded template, as long as the same cell
    positions carry the same fields.

    Known limitation: openpyxl does not reliably preserve embedded charts
    through a load/save round trip -- a workbook containing charts may
    lose them on export. Styles, merges, images, formulas, and layout are
    unaffected.
    """
    workbook = load_workbook(BytesIO(file_bytes), data_only=False, keep_vba=is_macro_enabled)
    ws = workbook.active

    anchor_for: dict[tuple[int, int], tuple[int, int]] = {}
    for merged_range in ws.merged_cells.ranges:
        anchor = (merged_range.min_row - 1, merged_range.min_col - 1)
        for r in range(merged_range.min_row - 1, merged_range.max_row):
            for c in range(merged_range.min_col - 1, merged_range.max_col):
                anchor_for[(r, c)] = anchor

    def _set(row0, col0, value):
        anchor_row0, anchor_col0 = anchor_for.get((row0, col0), (row0, col0))
        ws.cell(row=anchor_row0 + 1, column=anchor_col0 + 1).value = value

    for (r0, c0), key in HEADER_CELL_MAP.items():
        _, field = key.split(":", 1)
        _set(r0, c0, mold.get(field))

    for (r0, c0), key in EXTENDED_CELL_MAP.items():
        _, field = key.split(":", 1)
        _set(r0, c0, _extended_value(extended_fields, field))

    for (r0, c0), key in PARAMETER_CELL_MAP.items():
        _, tag = key.split(":", 1)
        _set(r0, c0, _tag_value(parameters_by_tag, tag))

    for offset, col0 in enumerate(_HOT_RUNNER_COLS):
        _, field = f"ext:hot_runner_t{offset + 1}".split(":", 1)
        _set(_HOT_RUNNER_ROW, col0, _extended_value(extended_fields, field))

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer