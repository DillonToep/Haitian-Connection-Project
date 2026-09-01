"""Reverses the cell maps in export_xlsx.py so a filled-in 试模成型参数表
can be read back into structured data for import.

Three source formats are supported, all mapped onto the same
HEADER_CELL_MAP / EXTENDED_CELL_MAP / PARAMETER_CELL_MAP /
_HOT_RUNNER_ROW / _HOT_RUNNER_COLS / _MERGES cell positions defined in
export_xlsx.py (so the maps only need to be maintained in one place):

- .xlsx / .xlsm -- read via openpyxl, using the workbook's own merged
  cells (a value only lives on a merge's top-left cell).
- .xls (Excel 97-2003 / legacy BIFF) -- read via xlrd, same idea but
  using xlrd's merged_cells info.
- .csv -- read as a plain row/column grid via the stdlib csv module.
  CSV has no merged-cell concept, so every mapped cell must contain its
  own value directly (i.e. a CSV exported from a merged sheet will only
  have the value in the top-left cell of each former merge anyway,
  which is exactly what these cell maps expect).

Only cells listed in the maps are read; anything else on the sheet is
ignored, same scope limitation the export side documents.
"""
from io import BytesIO, StringIO
import csv
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from .export_xlsx import (
    HEADER_CELL_MAP,
    EXTENDED_CELL_MAP,
    PARAMETER_CELL_MAP,
    _HOT_RUNNER_ROW,
    _HOT_RUNNER_COLS,
    _MERGES,
)
from .label_scan_xlsx import (
    build_resolved_grid,
    build_resolved_grid_xlrd,
    scan_all_blocks,
)
from .extended_field_scan import scan_extended_fields, scan_header_overrides

# Anchor-cell lookup for the *built-in* xlsx template's merges -- only
# used by the xlsx reader (openpyxl only stores a value on the top-left
# cell of a merged range; every other cell in that range reads back as
# None on a workbook built from our own template).
_ANCHOR_FOR = {}
for _r0, _r1, _c0, _c1 in _MERGES:
    for _rr in range(_r0, _r1):
        for _cc in range(_c0, _c1):
            _ANCHOR_FOR[(_rr, _cc)] = (_r0, _c0)


def _normalize_numeric_string(value):
    """Canonical form of a numeric value: a plain integer string ("25")
    when the value has no fractional part, otherwise a plain decimal
    string ("25.8") with no trailing zeros and no scientific notation.
    Never rounds/truncates a genuine decimal down to a whole number --
    a sheet value like 25.8 must stay 25.8, not become 26. Non-numeric
    values (enum codes, free text) pass through unchanged."""
    if value is None:
        return value
    candidate = value.strip() if isinstance(value, str) else value
    if candidate == "":
        return value
    try:
        number = Decimal(str(candidate))
    except (InvalidOperation, ValueError, TypeError):
        return str(value).strip() if isinstance(value, str) else value

    if number == number.to_integral_value():
        return str(int(number))

    text = format(number.normalize(), "f")
    return text


def _clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
    return value if value not in (None, "") else None


def _scan_parameters_xlsx(file_bytes: bytes) -> dict:
    """Label-driven parameter scan for .xlsx/.xlsm -- replaces reading
    PARAMETER_CELL_MAP by fixed coordinate. See label_scan_xlsx.py for
    why: an uploaded workbook's real row/column layout can differ from
    the built-in template's, and fixed coordinates silently misread the
    wrong cell (or a merged banner cell) when it does."""
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    worksheet = workbook.active
    grid = build_resolved_grid(worksheet)
    return scan_all_blocks(grid)


def _scan_parameters_xls(file_bytes: bytes) -> dict:
    try:
        import xlrd
    except ImportError as error:
        raise RuntimeError(
            "服务器缺少 xlrd 库，无法解析 .xls 文件（请安装 xlrd 或改用 .xlsx/.csv）"
        ) from error
    try:
        workbook = xlrd.open_workbook(file_contents=file_bytes, formatting_info=True)
    except Exception:
        workbook = xlrd.open_workbook(file_contents=file_bytes)
    sheet = workbook.sheet_by_index(0)
    grid = build_resolved_grid_xlrd(sheet)
    return scan_all_blocks(grid)


def _make_xlsx_reader(file_bytes: bytes):
    """.xlsx / .xlsm -- openpyxl. Resolves each mapped cell against the
    UPLOADED workbook's own merged ranges first (same approach the .xls
    reader below uses), falling back to the built-in template's static
    merge map (_ANCHOR_FOR) for any cell the uploaded file doesn't merge
    itself. The fallback keeps this working exactly as before for a
    workbook whose merges happen to match the built-in template, while
    the real-merges-first lookup fixes cells whose merge boundaries in
    the actual uploaded file differ from that static assumption (e.g. a
    template that was hand-edited, or one whose merges don't line up
    1:1 with ours) -- previously a mismatch here silently read back
    blank instead of resolving to wherever the value actually lives."""
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    worksheet = workbook.active

    anchor_for = dict(_ANCHOR_FOR)
    for merged_range in worksheet.merged_cells.ranges:
        anchor = (merged_range.min_row - 1, merged_range.min_col - 1)
        for r in range(merged_range.min_row - 1, merged_range.max_row):
            for c in range(merged_range.min_col - 1, merged_range.max_col):
                anchor_for[(r, c)] = anchor

    def _read(row0: int, col0: int):
        anchor_row0, anchor_col0 = anchor_for.get((row0, col0), (row0, col0))
        value = worksheet.cell(row=anchor_row0 + 1, column=anchor_col0 + 1).value
        return _clean(value)

    return _read


def _make_xls_reader(file_bytes: bytes):
    """Legacy .xls (Excel 97-2003 / BIFF), via xlrd. openpyxl cannot open
    this format, and xlrd 2.x dropped xlsx support entirely, so xlrd is
    used only for this one path."""
    try:
        import xlrd
    except ImportError as error:
        raise RuntimeError(
            "服务器缺少 xlrd 库，无法解析 .xls 文件（请安装 xlrd 或改用 .xlsx/.csv）"
        ) from error

    # formatting_info=True is required for xlrd to populate
    # sheet.merged_cells at all -- without it the attribute is always an
    # empty list regardless of what the file actually contains, which
    # silently defeated the merge-anchor resolution below for every cell
    # that isn't itself a merge's top-left corner (several of the
    # PARAMETER_CELL_MAP / EXTENDED_CELL_MAP entries aren't). A handful of
    # malformed/legacy .xls files raise while xlrd parses the formatting
    # records this needs, so fall back to the old (merge-blind) behavior
    # for those rather than failing the import outright.
    try:
        workbook = xlrd.open_workbook(file_contents=file_bytes, formatting_info=True)
    except Exception:
        workbook = xlrd.open_workbook(file_contents=file_bytes)
    sheet = workbook.sheet_by_index(0)

    # Build an anchor map from THIS file's own merged cells (not the
    # static template map above) -- an uploaded .xls may not be merged
    # identically to the generated xlsx template.
    anchor_for = {}
    for rlo, rhi, clo, chi in getattr(sheet, "merged_cells", []):
        for r in range(rlo, rhi):
            for c in range(clo, chi):
                anchor_for[(r, c)] = (rlo, clo)

    def _read(row0: int, col0: int):
        anchor_row0, anchor_col0 = anchor_for.get((row0, col0), (row0, col0))
        if anchor_row0 >= sheet.nrows or anchor_col0 >= sheet.ncols:
            return None
        value = sheet.cell_value(anchor_row0, anchor_col0)
        return _clean(value)

    return _read


def _make_csv_reader(file_bytes: bytes):
    """.csv -- plain row/column grid, no merged-cell concept. A mapped
    cell must carry its own value directly."""
    text = file_bytes.decode("utf-8-sig", errors="replace")
    rows = list(csv.reader(StringIO(text)))

    def _read(row0: int, col0: int):
        if row0 >= len(rows) or col0 >= len(rows[row0]):
            return None
        return _clean(rows[row0][col0])

    return _read


def _reader_for(file_bytes: bytes, filename: str):
    ext = (filename or "").lower().rsplit(".", 1)[-1] if "." in (filename or "") else ""
    if ext == "xls":
        return _make_xls_reader(file_bytes)
    if ext == "csv":
        return _make_csv_reader(file_bytes)
    return _make_xlsx_reader(file_bytes)  # .xlsx / .xlsm (default/fallback)


def parse_trial_parameter_workbook(file_bytes: bytes, filename: str = "") -> dict:
    """Returns
    {
        "header": {"mold_code": ..., "mold_name": ..., "cavities": ...},
        "extended": {key: value, ...},
        "parameters": {tag: normalized_value_string, ...},
    }
    Every sub-dict only contains keys the sheet actually had a
    non-blank value for -- a blank cell means "leave whatever is
    already saved alone", never "clear it".

    `filename` picks the parser: .xls -> xlrd, .csv -> csv module,
    anything else (.xlsx/.xlsm, or unspecified) -> openpyxl.
    """
    read_cell = _reader_for(file_bytes, filename)
    ext = (filename or "").lower().rsplit(".", 1)[-1] if "." in (filename or "") else ""

    # ---- build a resolved (merge-expanded) grid up front, when possible.
    # This is what both the parameter label-scan and the extended-field
    # label-scan below run against. Not attempted for .csv -- no merge
    # information exists to resolve a label-driven grid against there.
    grid = None
    try:
        if ext == "xls":
            import xlrd
            try:
                workbook = xlrd.open_workbook(file_contents=file_bytes, formatting_info=True)
            except Exception:
                workbook = xlrd.open_workbook(file_contents=file_bytes)
            sheet = workbook.sheet_by_index(0)
            grid = build_resolved_grid_xlrd(sheet)
        elif ext != "csv":
            workbook = load_workbook(BytesIO(file_bytes), data_only=True)
            grid = build_resolved_grid(workbook.active)
    except Exception:
        grid = None

    # Detect the newer 注塑成型条件参数表 layout: its header row always
    # carries a "客户名称" label, which never appears on the older
    # 试模成型参数表 layout HEADER_CELL_MAP/EXTENDED_CELL_MAP were built
    # for. On THAT older layout's fixed coordinates, a 注塑成型条件参数表
    # file reads back garbage (a value from the WRONG field entirely --
    # e.g. 机台厂商 pulling from 烘料时间's cell) rather than nothing, so
    # those fixed-coordinate maps must be skipped outright for this
    # layout rather than merely supplemented -- the label-driven scan in
    # extended_field_scan.py is trusted instead, since it locates each
    # value by its own label text rather than assuming a coordinate.
    is_new_layout = False
    if grid is not None:
        for row in grid:
            for value in row:
                if isinstance(value, str) and "客户名称" in value:
                    is_new_layout = True
                    break
            if is_new_layout:
                break

    header: dict = {}
    extended: dict = {}

    if not is_new_layout:
        for (row0, col0), key in HEADER_CELL_MAP.items():
            _, field = key.split(":", 1)
            value = read_cell(row0, col0)
            if value is not None:
                header[field] = value

        for (row0, col0), key in EXTENDED_CELL_MAP.items():
            _, field = key.split(":", 1)
            value = read_cell(row0, col0)
            if value is not None:
                extended[field] = value

        for offset, col0 in enumerate(_HOT_RUNNER_COLS):
            value = read_cell(_HOT_RUNNER_ROW, col0)
            if value is not None:
                extended[f"hot_runner_t{offset + 1}"] = value

    if grid is not None:
        extended.update(scan_extended_fields(grid))
        header.update(scan_header_overrides(grid))

    # ---- parameters: label-driven scan for .xlsx/.xlsm/.xls, since this
    # is the path that broke against a real-world layout drift (extra
    # inserted row shifted every fixed coordinate below it). .csv has no
    # merged-cell/label-search story worth building (a CSV export of a
    # merged sheet only carries a value in each former merge's top-left
    # cell anyway), so it stays on the fixed-coordinate reader.
    if ext == "xls":
        raw_parameters = _scan_parameters_xls(file_bytes)
    elif ext == "csv":
        raw_parameters = None
    else:
        raw_parameters = _scan_parameters_xlsx(file_bytes)

    parameters: dict = {}
    if raw_parameters is not None:
        for tag, value in raw_parameters.items():
            if value not in (None, ""):
                parameters[tag] = _normalize_numeric_string(value)
    else:
        for (row0, col0), key in PARAMETER_CELL_MAP.items():
            _, tag = key.split(":", 1)
            value = read_cell(row0, col0)
            if value is not None:
                parameters[tag] = _normalize_numeric_string(value)

    return {"header": header, "extended": extended, "parameters": parameters}